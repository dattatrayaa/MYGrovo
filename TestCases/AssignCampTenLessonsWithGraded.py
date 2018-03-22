'''
Created on 16-Mar-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearner import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from CreateTrackComman import CreateTrackComman


class AssignCampTenLessonsWithGraded:

    def campaignToAssignTenLesson(self,campaignTitle,campDescription,lessonName1,lessonName2,lessonName3,lessonName4,lessonName5,
                                            lessonName6,lessonName7,lessonName8,lessonName9,lessonName10,nameOFuser,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
       
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName1)
        elements.waitUntilSearchedLessonDisplayed(lessonName1)
        elements.selectSearchedLesson(lessonName1)
        
        elements.searchLesson(lessonName2)
        elements.waitUntilSearchedLessonDisplayed(lessonName2)
        elements.selectSearchedLesson(lessonName2)
        
        elements.searchLesson(lessonName3)
        elements.waitUntilSearchedLessonDisplayed(lessonName3)
        elements.selectSearchedLesson(lessonName3)
        
        elements.searchLesson(lessonName4)
        elements.waitUntilSearchedLessonDisplayed(lessonName4)
        elements.selectSearchedLesson(lessonName4)
        
        
        elements.searchLesson(lessonName5)
        elements.waitUntilSearchedLessonDisplayed(lessonName5)
        elements.selectSearchedLesson(lessonName5)
        
        
        elements.searchLesson(lessonName6)
        elements.waitUntilSearchedLessonDisplayed(lessonName6)
        elements.selectSearchedLesson(lessonName6)
        
        elements.searchLesson(lessonName7)
        elements.waitUntilSearchedLessonDisplayed(lessonName7)
        elements.selectSearchedLesson(lessonName7)
        
        elements.searchLesson(lessonName8)
        elements.waitUntilSearchedLessonDisplayed(lessonName8)
        elements.selectSearchedLesson(lessonName8)
        
        elements.searchLesson(lessonName9)
        elements.waitUntilSearchedLessonDisplayed(lessonName9)
        elements.selectSearchedLesson(lessonName9)
        
        elements.searchLesson(lessonName10)
        elements.waitUntilSearchedLessonDisplayed(lessonName10)
        elements.selectSearchedLesson(lessonName10)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
       
        time.sleep(2)
        wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//li/div[2]/div/h4/div/span"))) 
        
        titles=driver.find_elements_by_xpath("//li/div[2]/div/h4/div/span")
        count =len(titles)
        
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added all lesson's is displayed in Grid"
        
        
        i=1
        for count in (0,count):
           
            if driver.find_element_by_xpath("//li["+str(i)+"]/div[2]/div/h4/div/span").is_displayed():
                real=driver.find_element_by_xpath("//li["+str(i)+"]/div[2]/div/h4/div/span").text
                
                print "Lesson '"+real+"' is displayed"
                i=i+1
            else:
                print "Lesson is not displayed"
                raise Exception
            
        print "All lessons displayed in grid"
        
        
        
        
        
        time.sleep(2)
        
        
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        time.sleep(3)
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "user is not displayed in grid"
            raise Exception        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
    
    
    
    def assignCampWithTenLessonsGraded(self):
        
        book=xlrd.open_workbook(os.path.join('TestCases/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(108,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(109,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(129,1)
        nameOFuser = cell1.value
        
        cell1 = first_sheet.cell(110,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(111,1)
        numberOfAttempts = cell1.value
        
        
        
        #lesson
        cell2 = first_sheet.cell(112,1)
        lessonName1= cell2.value
        
        
        cell2 = first_sheet.cell(113,1)
        lessonName2= cell2.value
        
        cell2 = first_sheet.cell(114,1)
        lessonName3= cell2.value
        
        cell2 = first_sheet.cell(115,1)
        lessonName4= cell2.value
        
        cell2 = first_sheet.cell(116,1)
        lessonName5= cell2.value
        
        cell2 = first_sheet.cell(117,1)
        lessonName6= cell2.value
        
        cell2 = first_sheet.cell(118,1)
        lessonName7= cell2.value
        
        cell2 = first_sheet.cell(119,1)
        lessonName8= cell2.value
        
        cell2 = first_sheet.cell(120,1)
        lessonName9= cell2.value
        
        cell2 = first_sheet.cell(121,1)
        lessonName10= cell2.value
        
        
        cell2 = first_sheet.cell(122,1)
        textCard= cell2.value
        
        cell2 = first_sheet.cell(123,1)
        videoPath= cell2.value
        
        cell2 = first_sheet.cell(124,1)
        timeToUploadVideo= cell2.value
        
        
        cell2 = first_sheet.cell(125,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(126,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(127,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner
        cell = first_sheet.cell(129,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(130,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(131,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(132,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(133,1)
        Password = cell.value
        
        cell = first_sheet.cell(134,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(135,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=130, column=2).value = FirstNameUpdated
        sheet.cell(row=131, column=2).value = LastNameUpdated
        sheet.cell(row=132, column=2).value = EmailIdUpdated
        sheet.cell(row=133, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            #lt=CreateLearner()
            #lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, role, NewPassword, url, username, password)
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithText(lessonName1, textCard)
            ot.lessonWithText(lessonName2, textCard)
            ot.lessonWithText(lessonName3, textCard)
            
            ot.lessonWithVideo(lessonName4, videoPath, timeToUploadVideo)
            ot.lessonWithVideo(lessonName5, videoPath, timeToUploadVideo)
            ot.lessonWithVideo(lessonName6, videoPath, timeToUploadVideo)
            
            ot.lessonWithQuestion(lessonName7, questionCard, ans1, ans2)
            ot.lessonWithQuestion(lessonName8, questionCard, ans1, ans2)
            ot.lessonWithQuestion(lessonName9, questionCard, ans1, ans2)
            ot.lessonWithQuestion(lessonName10, questionCard, ans1, ans2)
            
            
            
            print "\nCreating Campaign\n"
            jk=AssignCampTenLessonsWithGraded()
            jk.campaignToAssignTenLesson(campaignTitle, campDescription, lessonName1, lessonName2, lessonName3, lessonName4, lessonName5, lessonName6, lessonName7, lessonName8, lessonName9, lessonName10, nameOFuser, minPassingScore, numberOfAttempts)
            
            
            print "\n----Text Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            #if any alert box occurs it will accept the alert
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            
            
            

    
