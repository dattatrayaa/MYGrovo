'''
Created on 16-Mar-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from CreateTrackComman import CreateTrackComman


class AssignCampNewHireTriggeredOneTrackAndOneLesson:
    
    
    def campaignTriggeredNewHireOneTrackOneLesson(self,campaignTitle,campDescription,lessonName,trackName,actualDue):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        #Adding track
        print "Adding Track"
        elements.addTrackButton()
        print "Searching Track"
        elements.searchTracksAndSelect(trackName)
        print "Track selected...Clicking on Add to campaign button"
        elements.addToCampaignTrack()
        print "Track Added to Campaign"
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        time.sleep(2)
        print "\nVerifying Added Track is displayed in Grid"
        if trackName in elements.firstTrackInGrid():
            print "Track displayed in grid"
        else:
            print "Track not displayed in Grid"
            raise Exception
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        print "Clicking on Triggered radio button"
        
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section/div/div[2]/label/span[2]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[1]/label/span[2]")))
        
        print "Clicked on Triggered"
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        print "Checking Due displayed"
        
        due=driver.find_element_by_xpath(".//*[@id='trigger-assignment-duration']").get_attribute("value")
        
        if due==actualDue:
            print "Due '"+due+"' is displayed"
        else:
            print "Due displayed is not correct"
        
        
        saveTrigger=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/div/div[2]/button")))
        saveTrigger.click()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[2]/button[1]")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[2]/button[1]").click()
        
        print "Clicked on Yes,Save button from pop up"
        newhier=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[3]/table/tbody/tr[1]/td[1]/div/span")))
        
        print "Checking for In Campaign details page Trigger is displayed"
        if newhier.text=="New hire onboarding":
            print "Trigger is displayed in Campaign details page"
        else:
            print "Trigger is not displayed"
            raise Exception
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
    
    
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
        
        
        
    def assignCampNewHireTriggeredOneTrackOneLesson(self):
        
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(139,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(140,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(141,1)
        actualDue = cell1.value
        
        
        cell1 = first_sheet.cell(142,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(143,1)
        textCard= cell2.value
        
        
        #track
        cell2 = first_sheet.cell(145,1)
        titleOfTrack= cell2.value
        
        cell2 = first_sheet.cell(146,1)
        Imagefilepath= cell2.value
        
        cell2 = first_sheet.cell(147,1)
        description= cell2.value
        
        cell2 = first_sheet.cell(148,1)
        tagName= cell2.value
        
        cell2 = first_sheet.cell(149,1)
        expectedSuccessText= cell2.value
        
        
        try:
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithText(lessonName, textCard)
            
            
            print "\nCreating Track\n"
            yl=CreateTrackComman()
            yl.createTrack(titleOfTrack, Imagefilepath, description, tagName, lessonName, expectedSuccessText)
            
            print "\nCreating Campaign\n"
            lk=AssignCampNewHireTriggeredOneTrackAndOneLesson()
            lk.campaignTriggeredNewHireOneTrackOneLesson(campaignTitle, campDescription, lessonName, titleOfTrack, actualDue)
            
            print "\n----Text Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            #if any alert box occurs it will accept the alert
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            
            
